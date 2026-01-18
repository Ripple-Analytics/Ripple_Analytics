#!/usr/bin/env python3
"""
Excel Export Builder
Generates comprehensive Excel reports from the Mental Models System.
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import psycopg2

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from config.settings import settings


class ExcelReportBuilder:
    """Builds comprehensive Excel reports from the Mental Models database."""
    
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = output_dir or settings.export.output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.conn = None
        self.workbook = None
    
    def connect(self):
        """Connect to the database."""
        self.conn = psycopg2.connect(
            dbname=settings.database.name,
            user=settings.database.user,
            host=settings.database.host,
            port=settings.database.port,
            password=settings.database.password or None
        )
    
    def close(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()
    
    def _apply_header_style(self, ws, row_num: int = 1):
        """Apply header styling to a row."""
        for cell in ws[row_num]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _query_to_dataframe(self, query: str) -> pd.DataFrame:
        """Execute query and return DataFrame."""
        return pd.read_sql_query(query, self.conn)
    
    def build_mental_models_sheet(self, ws):
        """Build the Mental Models overview sheet."""
        ws.title = "Mental Models"
        
        df = self._query_to_dataframe("""
            SELECT 
                name AS "Model Name",
                category AS "Category",
                description AS "Description",
                originator AS "Originator",
                modern_synthesizer AS "Modern Synthesizer",
                lindy_age_years AS "Lindy Age (Years)"
            FROM mental_models
            ORDER BY category, name
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        return len(df)
    
    def build_principles_sheet(self, ws):
        """Build the Thinker Principles sheet."""
        ws.title = "Thinker Principles"
        
        df = self._query_to_dataframe("""
            SELECT 
                thinker AS "Thinker",
                principle_name AS "Principle",
                principle_description AS "Description",
                category AS "Category",
                source AS "Source"
            FROM thinker_principles
            ORDER BY thinker, principle_name
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        return len(df)
    
    def build_case_studies_summary_sheet(self, ws):
        """Build the Case Studies summary sheet."""
        ws.title = "Case Studies Summary"
        
        df = self._query_to_dataframe("""
            SELECT 
                category AS "Category",
                COUNT(*) AS "Count",
                ROUND(AVG(severity)::numeric, 2) AS "Avg Severity",
                ROUND(AVG(financial_impact)::numeric, 0) AS "Avg Financial Impact",
                ROUND(AVG(models_involved)::numeric, 1) AS "Avg Models Involved",
                ROUND(AVG(lollapalooza_score)::numeric, 2) AS "Avg Lollapalooza Score"
            FROM case_studies
            GROUP BY category
            ORDER BY COUNT(*) DESC
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        if len(df) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Cases by Category"
            chart.y_axis.title = "Count"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1, max_col=2)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, "H2")
        
        return len(df)
    
    def build_decade_analysis_sheet(self, ws):
        """Build the Decade Analysis sheet."""
        ws.title = "Decade Analysis"
        
        df = self._query_to_dataframe("""
            SELECT 
                (EXTRACT(YEAR FROM date)::int / 10 * 10) AS "Decade",
                COUNT(*) AS "Case Count",
                ROUND(AVG(severity)::numeric, 2) AS "Avg Severity",
                ROUND(SUM(financial_impact)::numeric, 0) AS "Total Financial Impact"
            FROM case_studies
            WHERE date IS NOT NULL
            GROUP BY (EXTRACT(YEAR FROM date)::int / 10 * 10)
            ORDER BY "Decade"
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        if len(df) > 0:
            chart = LineChart()
            chart.title = "Cases Over Decades"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Decade"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "F2")
        
        return len(df)
    
    def build_region_analysis_sheet(self, ws):
        """Build the Region Analysis sheet."""
        ws.title = "Region Analysis"
        
        df = self._query_to_dataframe("""
            SELECT 
                region AS "Region",
                COUNT(*) AS "Case Count",
                ROUND(AVG(severity)::numeric, 2) AS "Avg Severity",
                ROUND(SUM(financial_impact)::numeric, 0) AS "Total Financial Impact"
            FROM case_studies
            GROUP BY region
            ORDER BY COUNT(*) DESC
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        if len(df) > 0:
            chart = PieChart()
            chart.title = "Cases by Region"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "F2")
        
        return len(df)
    
    def build_model_frequency_sheet(self, ws):
        """Build the Model Frequency analysis sheet."""
        ws.title = "Model Frequency"
        
        df = self._query_to_dataframe("""
            SELECT 
                model_name AS "Model",
                COUNT(*) AS "Occurrence Count",
                ROUND(AVG(effect_size)::numeric, 3) AS "Avg Effect Size",
                ROUND(STDDEV(effect_size)::numeric, 3) AS "Std Effect Size"
            FROM planck_matrix
            GROUP BY model_name
            ORDER BY COUNT(*) DESC
            LIMIT 50
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        return len(df)
    
    def build_lollapalooza_sheet(self, ws):
        """Build the Lollapalooza Patterns sheet."""
        ws.title = "Lollapalooza Patterns"
        
        df = self._query_to_dataframe("""
            SELECT 
                cs.name AS "Case Name",
                cs.category AS "Category",
                cs.severity AS "Severity",
                cs.models_involved AS "Models Involved",
                cs.lollapalooza_score AS "Lollapalooza Score",
                cs.financial_impact AS "Financial Impact"
            FROM case_studies cs
            WHERE cs.models_involved >= 5
            ORDER BY cs.lollapalooza_score DESC
            LIMIT 100
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        return len(df)
    
    def build_frameworks_sheet(self, ws):
        """Build the Frameworks (Thinkers) sheet."""
        ws.title = "Frameworks"
        
        df = self._query_to_dataframe("""
            SELECT 
                name AS "Thinker",
                description AS "Description",
                source AS "Primary Source"
            FROM frameworks
            ORDER BY name
        """)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)
        
        return len(df)
    
    def build_executive_summary_sheet(self, ws):
        """Build the Executive Summary sheet."""
        ws.title = "Executive Summary"
        
        stats = {}
        
        cur = self.conn.cursor()
        
        cur.execute("SELECT COUNT(*) FROM mental_models")
        stats['total_models'] = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM thinker_principles")
        stats['total_principles'] = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM case_studies")
        stats['total_cases'] = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM frameworks")
        stats['total_frameworks'] = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM planck_matrix")
        stats['total_planck_entries'] = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(DISTINCT category) FROM mental_models")
        stats['model_categories'] = cur.fetchone()[0]
        
        cur.close()
        
        ws['A1'] = "MENTAL MODELS SYSTEM - EXECUTIVE SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws['A5'] = "System Statistics"
        ws['A5'].font = Font(bold=True, size=12)
        
        summary_data = [
            ("Total Mental Models", stats['total_models']),
            ("Total Thinker Principles", stats['total_principles']),
            ("Total Case Studies", stats['total_cases']),
            ("Total Frameworks (Thinkers)", stats['total_frameworks']),
            ("Total Planck Matrix Entries", stats['total_planck_entries']),
            ("Model Categories", stats['model_categories']),
        ]
        
        for i, (label, value) in enumerate(summary_data, start=7):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        self._auto_adjust_columns(ws)
        
        return stats
    
    def generate_full_report(self, filename: Optional[str] = None) -> str:
        """Generate the complete Excel report."""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"mental_models_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        self.connect()
        self.workbook = Workbook()
        
        try:
            ws_summary = self.workbook.active
            self.build_executive_summary_sheet(ws_summary)
            
            ws_frameworks = self.workbook.create_sheet()
            self.build_frameworks_sheet(ws_frameworks)
            
            ws_models = self.workbook.create_sheet()
            self.build_mental_models_sheet(ws_models)
            
            ws_principles = self.workbook.create_sheet()
            self.build_principles_sheet(ws_principles)
            
            ws_cases = self.workbook.create_sheet()
            self.build_case_studies_summary_sheet(ws_cases)
            
            ws_decades = self.workbook.create_sheet()
            self.build_decade_analysis_sheet(ws_decades)
            
            ws_regions = self.workbook.create_sheet()
            self.build_region_analysis_sheet(ws_regions)
            
            ws_frequency = self.workbook.create_sheet()
            self.build_model_frequency_sheet(ws_frequency)
            
            ws_lollapalooza = self.workbook.create_sheet()
            self.build_lollapalooza_sheet(ws_lollapalooza)
            
            self.workbook.save(filepath)
            print(f"Report saved to: {filepath}")
            
        finally:
            self.close()
        
        return filepath


def export_to_csv(output_dir: Optional[str] = None) -> Dict[str, str]:
    """Export all tables to CSV files."""
    output_dir = output_dir or settings.export.output_dir
    os.makedirs(output_dir, exist_ok=True)
    
    conn = psycopg2.connect(
        dbname=settings.database.name,
        user=settings.database.user,
        host=settings.database.host,
        port=settings.database.port
    )
    
    tables = [
        'frameworks',
        'mental_models',
        'thinker_principles',
        'case_studies',
        'planck_matrix'
    ]
    
    exported = {}
    
    for table in tables:
        df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
        filepath = os.path.join(output_dir, f"{table}.csv")
        df.to_csv(filepath, index=False)
        exported[table] = filepath
        print(f"Exported {table} to {filepath}")
    
    conn.close()
    return exported


if __name__ == "__main__":
    builder = ExcelReportBuilder()
    report_path = builder.generate_full_report()
    print(f"Full report generated: {report_path}")
